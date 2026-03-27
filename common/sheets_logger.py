"""
매매 체결 시 Google Sheets / Excel 자동 기록 (OpenClaw gog 연동).

사용법:
  1. gog CLI: brew install steipete/tap/gogcli 후 gog auth add --services sheets
  2. 또는 gspread: GOOGLE_SHEETS_CREDENTIALS_JSON 경로 설정
  3. Excel: openpyxl 설치 후 append_trade(..., excel_path="./trades.xlsx")

환경변수:
  - GOOGLE_SHEET_ID: 시트 ID (필수)
  - GOOGLE_SHEET_TAB: 탭명 (기본 "거래기록")
  - GOG_KEYRING_PASSWORD: gog 사용 시 (openclaw-gog-secret)
"""
from __future__ import annotations

import json
import os
import shutil
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, List, Optional

# ── 환경 로드 ──────────────────────────────────────
try:
    from common.env_loader import load_env
    load_env()
except ImportError:
    pass

try:
    from common.logger import get_logger as _get_logger
    _log = _get_logger("sheets_logger")
except Exception:
    import logging
    _log = logging.getLogger("sheets_logger")

SHEET_ID = os.environ.get("GOOGLE_SHEET_ID", "")
SHEET_TAB = os.environ.get("GOOGLE_SHEET_TAB", "거래기록")
GOG_PASSWORD = os.environ.get("GOG_KEYRING_PASSWORD", "")

# Excel 기본 컬럼 헤더 (시트와 동일한 순서)
EXCEL_HEADERS = ["날짜", "마켓", "구분", "종목코드", "종목명", "가격", "수량", "수익률", "사유", "뉴스요약", "에이전트"]


def _escape(s: Any) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    return t.replace('"', '""').replace("\n", " ").replace("\r", "")


def _row(
    date: str,
    market: str,
    action: str,
    symbol: str,
    symbol_name: str,
    price: str,
    quantity: str,
    pnl_pct: str,
    reason: str,
    news_summary: str = "",
    agent: str = "",
) -> list[str]:
    return [
        date,
        market,
        action,
        symbol,
        symbol_name,
        price,
        quantity,
        pnl_pct,
        reason,
        news_summary,
        agent,
    ]


# ── gog 연동 ───────────────────────────────────────

def _append_via_gog(rows: list[list[str]]) -> bool:
    """gog CLI로 시트에 append."""
    workspace_dir = Path(__file__).resolve().parents[1]
    gog_path = workspace_dir / "gog-docker"

    if not gog_path.exists():
        gog_path = shutil.which("gog")
        if not gog_path:
            return False
    if not SHEET_ID:
        return False
    try:
        range_str = f"{SHEET_TAB}!A:K"
        values_json = json.dumps(rows, ensure_ascii=False)
        env = os.environ.copy()
        env["GOG_KEYRING_PASSWORD"] = GOG_PASSWORD
        result = subprocess.run(
            [
                str(gog_path),
                "sheets",
                "append",
                SHEET_ID,
                range_str,
                "--values-json",
                values_json,
                "--insert",
                "INSERT_ROWS",
            ],
            env=env,
            capture_output=True,
            text=True,
            timeout=15,
        )
        return result.returncode == 0
    except Exception:
        return False


def _append_via_gspread(rows: list[list[str]]) -> bool:
    """gspread로 시트에 append (fallback)."""
    creds_path = os.environ.get("GOOGLE_SHEETS_CREDENTIALS_JSON")
    if not creds_path or not Path(creds_path).exists():
        return False
    if not SHEET_ID:
        return False
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        gc = gspread.authorize(creds)
        sheet = gc.open_by_key(SHEET_ID).worksheet(SHEET_TAB)
        sheet.append_rows(rows, value_input_option="USER_ENTERED")
        return True
    except Exception:
        return False


# ── Excel(openpyxl) 연동 ─────────────────────────────

def _ensure_excel_headers(ws) -> None:
    """Excel 시트 1행이 비어있으면 헤더를 삽입한다."""
    if ws.max_row == 1 and all(ws.cell(1, c + 1).value is None for c in range(len(EXCEL_HEADERS))):
        for c, h in enumerate(EXCEL_HEADERS, 1):
            ws.cell(1, c, h)


def append_to_excel(rows: List[List[Any]], excel_path: str, tab: str = "거래기록") -> bool:
    """Excel 파일(.xlsx)에 행을 추가한다.

    Args:
        rows: 추가할 행 목록 (각 행은 EXCEL_HEADERS 순서)
        excel_path: .xlsx 파일 경로 (없으면 새로 생성)
        tab: 시트 탭 이름 (기본 "거래기록")
    """
    try:
        import openpyxl
    except ImportError:
        _log.warning("openpyxl 미설치 — pip install openpyxl")
        return False

    try:
        path = Path(excel_path)
        if path.exists():
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
            # 기본 시트 이름 변경
            if wb.active:
                wb.active.title = tab

        if tab in wb.sheetnames:
            ws = wb[tab]
        else:
            ws = wb.create_sheet(tab)

        _ensure_excel_headers(ws)

        for row in rows:
            ws.append(row)

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return True
    except Exception as e:
        _log.error(f"Excel append 실패: {e}")
        return False


def read_from_excel(excel_path: str, tab: str = "거래기록", skip_header: bool = True) -> List[List[Any]]:
    """Excel 파일(.xlsx)에서 데이터를 읽어 반환한다.

    Args:
        excel_path: .xlsx 파일 경로
        tab: 시트 탭 이름
        skip_header: True면 1행(헤더) 건너뜀

    Returns:
        list of rows (각 행은 list[Any])
    """
    try:
        import openpyxl
    except ImportError:
        _log.warning("openpyxl 미설치 — pip install openpyxl")
        return []

    try:
        path = Path(excel_path)
        if not path.exists():
            return []
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if tab not in wb.sheetnames:
            _log.warning(f"Excel 탭 없음: {tab}")
            return []
        ws = wb[tab]
        rows = list(ws.values)
        return rows[1:] if skip_header and rows else rows
    except Exception as e:
        _log.error(f"Excel read 실패: {e}")
        return []


def read_sheet_via_gspread(
    sheet_id: Optional[str] = None,
    tab: Optional[str] = None,
    skip_header: bool = True,
) -> List[List[Any]]:
    """gspread로 Google Sheets 데이터를 읽어 반환한다.

    Args:
        sheet_id: 구글 시트 ID (None이면 GOOGLE_SHEET_ID 환경변수)
        tab: 탭명 (None이면 GOOGLE_SHEET_TAB 환경변수)
        skip_header: True면 1행 건너뜀

    Returns:
        list of rows (각 행은 list[Any])
    """
    creds_path = os.environ.get("GOOGLE_SHEETS_CREDENTIALS_JSON")
    sid = sheet_id or SHEET_ID
    stab = tab or SHEET_TAB
    if not creds_path or not Path(creds_path).exists() or not sid:
        return []
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        gc = gspread.authorize(creds)
        ws = gc.open_by_key(sid).worksheet(stab)
        all_values = ws.get_all_values()
        return all_values[1:] if skip_header and all_values else all_values
    except Exception as e:
        _log.error(f"Sheets read 실패 ({stab}): {e}")
        return []


# ── 통합 append_trade ──────────────────────────────

def append_trade(
    market: str,
    action: str,
    symbol: str,
    price: float,
    quantity: float,
    pnl_pct: Optional[float] = None,
    reason: str = "",
    news_summary: str = "",
    symbol_name: str = "",
    agent: str = "",
    excel_path: Optional[str] = None,
) -> bool:
    """
    매매 체결 시 구글 시트 및/또는 Excel 파일에 한 행 추가.

    Args:
        market: "btc" | "kr" | "us"
        action: "매수" | "매도" | "손절" | "익절"
        symbol: "BTC" | "005930" | "AAPL" 등 종목코드
        symbol_name: "비트코인" | "삼성전자" | "애플" 등 종목명
        price: 체결가
        quantity: 수량
        pnl_pct: 수익률 (매도/손절/익절 시)
        reason: 진입/청산 근거
        news_summary: 당시 뉴스 요약 (선택)
        agent: 에이전트명 (자동 완성 가능)
        excel_path: Excel 파일 경로 (설정 시 Excel에도 기록)
    """
    if not SHEET_ID and not excel_path:
        return False

    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    # 가격 포맷팅
    if isinstance(price, (int, float)):
        if market.lower() == "btc":
            price_str = f"{price:,.0f}"
        elif market.lower() == "us":
            price_str = f"${price:,.2f}"
        else:
            price_str = f"{price:,.0f}"
    else:
        price_str = str(price)

    # 수량 포맷팅
    if isinstance(quantity, (int, float)):
        if market.lower() == "btc":
            qty_str = f"{quantity:.6f}"
        elif market.lower() == "us":
            qty_str = f"{quantity:.2f}"
        else:
            qty_str = f"{quantity:.0f}"
    else:
        qty_str = str(quantity)

    pnl_str = f"{pnl_pct:+.2f}%" if pnl_pct is not None else ""

    # 종목명 자동 완성
    if not symbol_name:
        symbol_names = {
            "BTC": "비트코인",
            "005930": "삼성전자",
            "000660": "SK하이닉스",
            "AAPL": "애플",
            "TSLA": "테슬라",
            "MSFT": "마이크로소프트",
            "NVDA": "엔비디아",
            "AMZN": "아마존",
            "GOOGL": "구글",
            "META": "메타",
        }
        symbol_name = symbol_names.get(symbol.upper(), symbol)

    if not agent:
        agent = f"{market}_agent"

    row = _row(
        date=date_str,
        market=market.upper(),
        action=action,
        symbol=symbol,
        symbol_name=symbol_name,
        price=price_str,
        quantity=qty_str,
        pnl_pct=pnl_str,
        reason=_escape(reason)[:200],
        news_summary=_escape(news_summary)[:150],
        agent=agent,
    )

    results = []

    # 1) Google Sheets (gog → gspread 순 시도)
    if SHEET_ID:
        if _append_via_gog([row]):
            results.append(True)
        elif _append_via_gspread([row]):
            results.append(True)
        else:
            results.append(False)

    # 2) Excel (선택적)
    if excel_path:
        # Excel에는 raw 숫자값 기록 (수식 활용 가능)
        excel_row = [
            date_str,
            market.upper(),
            action,
            symbol,
            symbol_name,
            price if isinstance(price, (int, float)) else price_str,
            quantity if isinstance(quantity, (int, float)) else qty_str,
            pnl_pct if pnl_pct is not None else "",
            _escape(reason)[:200],
            _escape(news_summary)[:150],
            agent,
        ]
        results.append(append_to_excel([excel_row], excel_path, tab=SHEET_TAB))

    return any(results)


def is_configured() -> bool:
    """구글 시트 또는 Excel 연동이 설정되어 있는지 확인."""
    workspace_dir = Path(__file__).resolve().parents[1]
    gog_docker_path = workspace_dir / "gog-docker"
    has_gog = gog_docker_path.exists() or shutil.which("gog")
    sheets_ok = bool(SHEET_ID and (has_gog or os.environ.get("GOOGLE_SHEETS_CREDENTIALS_JSON")))
    excel_ok = bool(os.environ.get("EXCEL_TRADE_LOG_PATH"))
    return sheets_ok or excel_ok
